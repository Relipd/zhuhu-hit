# -*- coding: utf-8 -*-
"""知乎热榜跟进 - 填月度 Excel(不存在时自动建模板; 新日期 sheet 插到最前)。

用法:
  python fill_excel.py --root <工作根目录> --date 2026-08-08 [--xlsx <文件路径>]

读: <root>/raw/<date>/{hot.json, answers_summary.json, analysis.json}
约定: 一级行=问题, 二级行=回答(≤10条); 月份文件内按天分 sheet, 最新日期在前。
"""
import argparse, json, os, sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = ["层级", "问题序号", "排名", "问题标题", "原问题URL", "问题点赞数", "问题本质",
           "回答序号", "回答内容", "回答点赞数",
           "立场分析", "解决思路", "判断逻辑", "情绪倾向", "情绪判断", "备注"]
FONT = "微软雅黑"
HDR_FILL = PatternFill("solid", fgColor="2F5496")
HDR_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
Q_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="top")
WIDTHS = [7, 8, 6, 40, 36, 9, 40, 8, 50, 9, 26, 26, 26, 26, 10, 12]

DOC = [
    ("结构", "按天分页: 每天一个 sheet(YYYY-MM-DD), 最新日期排最前。每月一个 Excel 文件。"),
    ("一级行-问题", "层级=问题: 问题序号(1-20), 排名, 标题, 原问题URL, 问题点赞数(接口无则取最高赞回答并备注), 问题本质。"),
    ("二级行-回答", "层级=回答: 问题序号(归属), 回答序号(1-10), 回答内容(原文), 回答点赞数。"),
    ("回答四维分析", "立场分析/解决思路/判断逻辑/情绪倾向 基于原文归纳, 不编造; 情绪判断三选一 积极/中立/消极。"),
    ("示例行", "模板首次创建时含示例日 sheet, 正式抓取后由 fill_excel 替换。"),
    ("数据来源", "zhihu-cli: hot + search zhihu 多变体查询合并。详见 skill 文档。"),
]

def ensure_workbook(path, year, month):
    """无文件时创建月度模板(含说明 sheet, 无示例日)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 100
    for i, (a, b) in enumerate(DOC, 1):
        ca, cb = ws.cell(row=i, column=1, value=a), ws.cell(row=i, column=2, value=b)
        ca.font = Font(name=FONT, size=10, bold=True)
        cb.font = Font(name=FONT, size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True)
        if i == 1:
            for c in (ca, cb):
                c.fill, c.font = HDR_FILL, HDR_FONT
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"新建月度模板: {path}")

def style_row(ws, is_q):
    for c in ws[ws.max_row]:
        c.font = Font(name=FONT, size=10)
        c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if is_q:
            c.fill = Q_FILL
    for col in (2, 3, 6, 8, 10, 15):
        ws.cell(row=ws.max_row, column=col).alignment = CENTER

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=os.getcwd())
    ap.add_argument("--date", required=True, help="YYYY-MM-DD")
    ap.add_argument("--xlsx", default=None, help="Excel 路径(默认 <root>/跟进excel-YYYY-MM.xlsx)")
    args = ap.parse_args()

    day_dir = os.path.join(args.root, "raw", args.date)
    year, month = args.date.split("-")[0], args.date.split("-")[1]
    xlsx = args.xlsx or os.path.join(args.root, f"跟进excel-{year}-{month}.xlsx")
    if not os.path.exists(xlsx):
        ensure_workbook(xlsx, year, month)

    hot = json.load(open(os.path.join(day_dir, "hot.json"), encoding="utf-8-sig"))
    summary = json.load(open(os.path.join(day_dir, "answers_summary.json"), encoding="utf-8"))
    an = json.load(open(os.path.join(day_dir, "analysis.json"), encoding="utf-8"))
    ext_path = os.path.join(day_dir, "extension.json")
    ext = json.load(open(ext_path, encoding="utf-8")) if os.path.exists(ext_path) else None

    wb = load_workbook(xlsx)
    if args.date in wb.sheetnames:
        del wb[args.date]
    ws = wb.create_sheet(args.date, 0)
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for s in summary:
        rank, q = s["rank"], an[str(s["rank"])]
        top = s["answers"][0]["likes"] if s["answers"] else None
        ws.append(["问题", rank, rank, s["title"], s["url"], top, q["essence"],
                   None, None, None, None, None, None, None, None,
                   "问题点赞数=该问题最高赞回答" if top is not None else None])
        style_row(ws, True)
        ws.cell(row=ws.max_row, column=5).hyperlink = s["url"]
        for i, a in enumerate(s["answers"], 1):
            A = q["answers"][i - 1]
            note = {"truncated": "接口摘要，全文需登录网页查看", "summary": "接口摘要(全文抓取失败)",
                    "full": None}.get(a.get("content_status"))
            ws.append(["回答", rank, None, None, None, None, None, i, a["text"], a["likes"],
                       A["stance"], A["approach"], A["logic"], A["emotion"], A["judge"], note])
            style_row(ws, False)

    # 情绪判断列(O列)三值下拉, 标签化约束
    dv = DataValidation(type="list", formula1='"积极,中立,消极"', allow_blank=True,
                        showErrorMessage=True, errorTitle="情绪判断",
                        error="仅允许: 积极 / 中立 / 消极(由 Agent 阅读原文判断, 禁止脚本/程序判定)")
    ws.add_data_validation(dv)
    dv.add(f"O2:O{ws.max_row}")

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{ws.max_row}"

    # ---- 热点拓展板块(extension.json, 仅热榜前10) ----
    if ext:
        EXT_HEADERS = ["日期", "排名", "问题标题", "扩展类型", "扩展内容", "来源链接", "备注"]
        ext_sheet = wb.create_sheet("热点拓展") if "热点拓展" not in wb.sheetnames else wb["热点拓展"]
        # 保留其他日期的行
        old_rows = []
        if ext_sheet.max_row > 1:
            for r in ext_sheet.iter_rows(min_row=2, values_only=True):
                if r[0] != args.date:
                    old_rows.append(r)
        ext_sheet.delete_rows(2, ext_sheet.max_row)
        ext_sheet.append(EXT_HEADERS)
        for c in ext_sheet[1]:
            c.fill, c.font = HDR_FILL, HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        new_rows = []
        for rank_str in sorted(ext.keys(), key=lambda k: int(k)):
            e = ext[rank_str]
            for it in e.get("items", []):
                new_rows.append([args.date, int(rank_str), e["title"], it["type"], it["content"],
                                 it.get("url", ""), it.get("note", "")])
            if e.get("thinking"):
                new_rows.append([args.date, int(rank_str), e["title"], "思考过程", e["thinking"], "", ""])
        all_rows = new_rows + old_rows
        all_rows.sort(key=lambda r: (str(r[0]), r[1]), reverse=True)  # 最新日期在上
        for r in all_rows:
            ext_sheet.append(r)
            for c in ext_sheet[ext_sheet.max_row]:
                c.font = Font(name=FONT, size=10)
                c.border = BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
            u = ext_sheet.cell(row=ext_sheet.max_row, column=6)
            if r[5]:
                u.hyperlink = r[5]
        for i, w in enumerate([11, 6, 36, 10, 52, 36, 18], 1):
            ext_sheet.column_dimensions[get_column_letter(i)].width = w
        ext_sheet.freeze_panes = "A2"
        ext_sheet.auto_filter.ref = f"A1:G{ext_sheet.max_row}"
        wb.move_sheet("热点拓展", offset=len(wb.sheetnames) - 2)

    if "说明" in wb.sheetnames:
        wb.move_sheet("说明", offset=len(wb.sheetnames) - 1)
    wb.save(xlsx)
    print(f"saved: {xlsx} | sheet {args.date}: {len(summary)} 问题, "
          f"{sum(len(s['answers']) for s in summary)} 回答, 共 {ws.max_row - 1} 行"
          + (f" | 热点拓展: {len(new_rows) if ext else 0} 行" if ext else ""))

if __name__ == "__main__":
    main()
